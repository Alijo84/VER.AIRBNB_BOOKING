import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

# Funciones de verificación (debes definir estas funciones)
def revisar_airbnb(uploaded_file):
    archivo_excel_original = uploaded_file
    hoja_consolidados = "CONSOLIDADO BANCO"
    hoja_pagos_airbnb = "PAGO AIRBNB "
    hoja_smoobu_airbnb = "SMOOBU AIRBNB"
    nuevo_archivo_excel = "RESERVA_SMOOBOO_VERIFICADA_AIRBNB.xlsx"
    nueva_hoja_consolidados = "consolidados bancos verificada"
    nueva_hoja_pagos_airbnb = "PAGOS AIRBNB VERIFICADA"
    nueva_hoja_smoobu_airbnb = "SMOOBU AIRBNB VERIFICADA"

    # Leer y procesar el archivo original
    xls = pd.ExcelFile(archivo_excel_original)

    if hoja_consolidados not in xls.sheet_names:
        st.error(f"Error: La hoja '{hoja_consolidados}' no existe en el archivo original.")
        return

    if hoja_pagos_airbnb not in xls.sheet_names:
        st.error(f"Error: La hoja '{hoja_pagos_airbnb}' no existe en el archivo original.")
        return

    if hoja_smoobu_airbnb not in xls.sheet_names:
        st.error(f"Error: La hoja '{hoja_smoobu_airbnb}' no existe en el archivo original.")
        return

    # Cargar la hoja 'CONSOLIDADO BANCO'
    df_consolidados = pd.read_excel(archivo_excel_original, sheet_name=hoja_consolidados)

    # Filtrar solo las filas donde 'CONCEPTO' sea 'AIRBNB'
    df_consolidados_filtrados = df_consolidados[df_consolidados['CONCEPTO'] == 'AIRBNB']

    # Seleccionar solo las columnas 'BANCO', 'FECHA', 'MONTO' y 'CONCEPTO'
    df_consolidados_filtrados = df_consolidados_filtrados[['BANCO', 'FECHA', 'MONTO', 'CONCEPTO']]

    # Convertir la columna de fecha a tipo datetime
    df_consolidados_filtrados['FECHA'] = pd.to_datetime(df_consolidados_filtrados['FECHA'], dayfirst=True)

    # Ordenar el DataFrame por la columna de fecha en orden descendente
    df_consolidados_filtrados_sorted = df_consolidados_filtrados.sort_values(by='FECHA', ascending=False)

    # Formatear la columna de fecha para que aparezca sin la hora y en el orden día/mes/año
    df_consolidados_filtrados_sorted['FECHA'] = df_consolidados_filtrados_sorted['FECHA'].dt.strftime('%d/%m/%Y')

    # Crear una copia de df_consolidados_sorted para la hoja verificada
    df_consolidados_verificada = df_consolidados_filtrados_sorted.copy()

    # Leer la hoja 'PAGOS AIRBNB'
    df_pagos = pd.read_excel(archivo_excel_original, sheet_name=hoja_pagos_airbnb)

    # Convertir la columna de fecha a tipo datetime y formatear
    df_pagos['Fecha'] = pd.to_datetime(df_pagos['Fecha'], dayfirst=True)
    df_pagos['Fecha'] = df_pagos['Fecha'].dt.strftime('%d/%m/%Y')

    # Crear una copia de df_pagos para la hoja verificada
    df_pagos_verificada = df_pagos.copy()

    # Comparar y verificar pagos
    def verificar_pago(row):
        monto = row['Cobrado']
        tipo = row['Tipo']
        fecha = pd.to_datetime(row['Fecha'], dayfirst=True)
        if tipo.lower() == "payout":
            df_filtrado = df_consolidados_verificada[
                (pd.to_datetime(df_consolidados_verificada['FECHA'], dayfirst=True) >= fecha) &
                (df_consolidados_verificada['CONCEPTO'].str.contains('AIRBNB', case=False)) &
                (df_consolidados_verificada['MONTO'] == monto)
            ]
            if not df_filtrado.empty:
                return "PAGADO"
            else:
                return "NO PAGADO"
        return ""  # Devolver vacío si el tipo no es 'payout'

    # Aplicar la verificación y crear la columna 'PAGO DEFINITIVO'
    df_pagos_verificada['PAGO DEFINITIVO'] = df_pagos_verificada.apply(verificar_pago, axis=1)

    # Función para calcular la columna 'OBS'
    def calcular_obs(df):
        obs = [""] * len(df)  # Inicializa la columna OBS con valores vacíos

        for i, row in df.iterrows():
            if row['Tipo'].lower() == "payout" and pd.notna(row['Cobrado']):
                suma_importe = 0
                cobrado = row['Cobrado']

                # Iterar a partir de la siguiente fila
                for j in range(i + 1, len(df)):
                    if pd.isna(df.at[j, 'Importe']) or df.at[j, 'Tipo'].lower() != "reserva":
                        break
                    suma_importe += df.at[j, 'Importe']

                # Comparar la suma con el valor de 'Cobrado'
                if suma_importe == cobrado:
                    for j in range(i + 1, len(df)):
                        if df.at[j, 'Tipo'].lower() == "reserva":
                            obs[j] = "P" if row['PAGO DEFINITIVO'] == "PAGADO" else "NP"

        return obs

    # Aplicar la función para calcular 'OBS' y agregar la columna a 'df_pagos_verificada'
    df_pagos_verificada['OBS'] = calcular_obs(df_pagos_verificada)

    # Leer la hoja 'SMOOBU AIRBNB'
    df_smoobu_airbnb = pd.read_excel(archivo_excel_original, sheet_name=hoja_smoobu_airbnb)

    # Agregar la columna 'Pago Neto'
    df_smoobu_airbnb['Pago Neto'] = df_smoobu_airbnb['Precio'] - df_smoobu_airbnb['Comisión incluida']

    # Reordenar las columnas para que 'Pago Neto' esté al lado de 'Comisión incluida'
    cols = df_smoobu_airbnb.columns.tolist()
    index_comision = cols.index('Comisión incluida')
    cols.insert(index_comision + 1, cols.pop(cols.index('Pago Neto')))
    df_smoobu_airbnb = df_smoobu_airbnb[cols]

    # Crear una función para observar pagos y verificar las fechas y montos
    def observar_pago(row):
        reserva = row['reserva']
        fecha_llegada = pd.to_datetime(row['Llegada'], dayfirst=True)
        pago_neto = row['Pago Neto']
        
        df_filtrado = df_pagos_verificada[df_pagos_verificada['Código de confirmación'] == reserva]
        
        if not df_filtrado.empty:
            return "PAGADO"
        
        df_filtrado_consolidados = df_consolidados_verificada[
            (pd.to_datetime(df_consolidados_verificada['FECHA'], dayfirst=True) >= fecha_llegada) &
            (df_consolidados_verificada['MONTO'] == pago_neto)
        ]
        
        if not df_filtrado_consolidados.empty:
            return "PAGADO"
        
        return "NOS HAN TIMADO"

    # Aplicar la observación y crear la columna 'OBSERVACION'
    df_smoobu_airbnb['OBSERVACION'] = df_smoobu_airbnb.apply(observar_pago, axis=1)

    # Reordenar las columnas para que 'OBSERVACION' esté al lado de 'reserva'
    cols = df_smoobu_airbnb.columns.tolist()
    index_reserva = cols.index('reserva')
    cols.insert(index_reserva + 1, cols.pop(cols.index('OBSERVACION')))
    df_smoobu_airbnb = df_smoobu_airbnb[cols]

    # Escribir los DataFrames en un nuevo archivo Excel con las nuevas hojas
    with pd.ExcelWriter(nuevo_archivo_excel, engine='openpyxl') as writer:
        df_consolidados_verificada.to_excel(writer, sheet_name=nueva_hoja_consolidados, index=False)
        df_pagos_verificada.to_excel(writer, sheet_name=nueva_hoja_pagos_airbnb, index=False)
        df_smoobu_airbnb.to_excel(writer, sheet_name=nueva_hoja_smoobu_airbnb, index=False)

    # Cargar el archivo Excel para aplicar el formato condicional
    wb = load_workbook(nuevo_archivo_excel)

    # Aplicar formato condicional a la hoja 'PAGOS AIRBNB VERIFICADA'
    ws_pagos_airbnb = wb[nueva_hoja_pagos_airbnb]
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for row in ws_pagos_airbnb.iter_rows(min_row=2, min_col=df_pagos_verificada.columns.get_loc('PAGO DEFINITIVO') + 1,
                                         max_col=df_pagos_verificada.columns.get_loc('PAGO DEFINITIVO') + 1, max_row=ws_pagos_airbnb.max_row):
        for cell in row:
            if cell.value == "PAGADO":
                cell.fill = green_fill
            elif cell.value == "NO PAGADO":
                cell.fill = red_fill

    # Aplicar formato condicional a la hoja 'SMOOBU AIRBNB VERIFICADA'
    ws_smoobu_airbnb = wb[nueva_hoja_smoobu_airbnb]

    for row in ws_smoobu_airbnb.iter_rows(min_row=2, min_col=df_smoobu_airbnb.columns.get_loc('OBSERVACION') + 1,
                                          max_col=df_smoobu_airbnb.columns.get_loc('OBSERVACION') + 1, max_row=ws_smoobu_airbnb.max_row):
        for cell in row:
            if cell.value == "PAGADO":
                cell.fill = green_fill
            elif cell.value == "NOS HAN TIMADO":
                cell.fill = red_fill

    # **Nuevo**: Aplicar formato condicional a la columna 'estado' en la hoja 'SMOOBU AIRBNB VERIFICADA'
    estado_col_idx = df_smoobu_airbnb.columns.get_loc('estado') + 1
    for row in ws_smoobu_airbnb.iter_rows(min_row=2, min_col=estado_col_idx, max_col=estado_col_idx, max_row=ws_smoobu_airbnb.max_row):
        for cell in row:
            if cell.value == "Cancelado":
                cell.fill = red_fill

    # Guardar el archivo con los cambios
    wb.save(nuevo_archivo_excel)

    st.success(f"El nuevo archivo '{nuevo_archivo_excel}' ha sido creado con las hojas '{nueva_hoja_consolidados}', '{nueva_hoja_pagos_airbnb}', y '{nueva_hoja_smoobu_airbnb}'.")

    return nuevo_archivo_excel


def revisar_booking(uploaded_file):
    archivo_excel_original = uploaded_file
    hoja_consolidados = "CONSOLIDADO BANCO"
    hoja_pagos_booking = "PAGO BOOKING"
    hoja_smoobu_booking = "SMOOBU BOOKING"
    nuevo_archivo_excel = "RESERVA_SMOOBOO_VERIFICADA_BOOKING.xlsx"
    nueva_hoja_consolidados = "consolidados bancos verificada"
    nueva_hoja_pagos_booking = "PAGOS BOOKING VERIFICADA"
    nueva_hoja_smoobu_booking = "SMOOBU BOOKING VERIFICADA"

    # Leer y procesar el archivo original
    xls = pd.ExcelFile(archivo_excel_original)

    if hoja_consolidados not in xls.sheet_names:
        st.error(f"Error: La hoja '{hoja_consolidados}' no existe en el archivo original.")
        return

    if hoja_pagos_booking not in xls.sheet_names:
        st.error(f"Error: La hoja '{hoja_pagos_booking}' no existe en el archivo original.")
        return

    if hoja_smoobu_booking not in xls.sheet_names:
        st.error(f"Error: La hoja '{hoja_smoobu_booking}' no existe en el archivo original.")
        return

    # Cargar la hoja 'CONSOLIDADO BANCO'
    df_consolidados = pd.read_excel(archivo_excel_original, sheet_name=hoja_consolidados)

    # Filtrar solo las filas donde 'CONCEPTO' sea 'BOOKING'
    df_consolidados_filtrados = df_consolidados[df_consolidados['CONCEPTO'] == 'BOOKING']

    # Seleccionar solo las columnas 'BANCO', 'FECHA', 'MONTO' y 'CONCEPTO'
    df_consolidados_filtrados = df_consolidados_filtrados[['BANCO', 'FECHA', 'MONTO', 'CONCEPTO']]

    # Convertir la columna de fecha a tipo datetime
    df_consolidados_filtrados['FECHA'] = pd.to_datetime(df_consolidados_filtrados['FECHA'], dayfirst=True)

    # Ordenar el DataFrame por la columna de fecha en orden descendente
    df_consolidados_filtrados_sorted = df_consolidados_filtrados.sort_values(by='FECHA', ascending=False)

    # Formatear la columna de fecha para que aparezca sin la hora y en el orden día/mes/año
    df_consolidados_filtrados_sorted['FECHA'] = df_consolidados_filtrados_sorted['FECHA'].dt.strftime('%d/%m/%Y')

    # Leer la hoja 'PAGOS BOOKING'
    df_pagos = pd.read_excel(archivo_excel_original, sheet_name=hoja_pagos_booking)

    # Filtrar valores no válidos en la columna 'Payout date'
    df_pagos = df_pagos[pd.to_datetime(df_pagos['Payout date'], errors='coerce').notna()]

    # Convertir la columna de fecha a tipo datetime y formatear
    df_pagos['Payout date'] = pd.to_datetime(df_pagos['Payout date'], dayfirst=True)
    df_pagos['Payout date'] = df_pagos['Payout date'].dt.strftime('%d/%m/%Y')

    # Asegurarse de que la columna 'Net' tenga solo valores numéricos y redondear a dos decimales
    df_pagos['Net'] = pd.to_numeric(df_pagos['Net'], errors='coerce').round(2)

    # Sumar las cantidades en la columna 'Net' cuando el 'Payout ID' es igual
    df_pagos['TOTAL DEPOSITADO'] = df_pagos.groupby('Payout ID')['Net'].transform('sum')

    # Redondear 'TOTAL DEPOSITADO' a dos decimales
    df_pagos['TOTAL DEPOSITADO'] = df_pagos['TOTAL DEPOSITADO'].round(2)

    # Evitar que el total depositado se repita
    df_pagos['TOTAL DEPOSITADO'] = df_pagos.groupby('Payout ID')['TOTAL DEPOSITADO'].transform(lambda x: x.where(x.index == x.index.min(), ""))

    # Crear una función para observar pagos
    def observar_pago(row):
        total_depositado = row['TOTAL DEPOSITADO']
        if total_depositado == "":
            return ""  # Dejar celda vacía si TOTAL DEPOSITADO está vacío
        payout_date = pd.to_datetime(row['Payout date'], format='%d/%m/%Y')
        df_filtrado = df_consolidados_filtrados_sorted[
            (df_consolidados_filtrados_sorted['MONTO'] == total_depositado) &
            (pd.to_datetime(df_consolidados_filtrados_sorted['FECHA'], format='%d/%m/%Y') >= payout_date)
        ]
        if not df_filtrado.empty:
            return "PAGADO"
        return "NO PAGADO"

    # Aplicar la observación y crear la columna 'OBSERVACION'
    df_pagos['OBSERVACION'] = df_pagos.apply(observar_pago, axis=1)

    # Reordenar las columnas para que 'OBSERVACION' esté justo después de 'TOTAL DEPOSITADO'
    cols = df_pagos.columns.tolist()
    index_total_depositado = cols.index('TOTAL DEPOSITADO')
    cols.insert(index_total_depositado + 1, cols.pop(cols.index('OBSERVACION')))
    df_pagos = df_pagos[cols]

    # Crear una copia de df_pagos para la hoja verificada
    df_pagos_verificada = df_pagos.copy()

    # Añadir columna 'OBS' al lado derecho de 'OBSERVACION' en df_pagos_verificada
    df_pagos_verificada['OBS'] = ""

    # Agrupar por 'Payout ID' y sumar los valores de 'Net'
    suma_nets = df_pagos_verificada.groupby('Payout ID')['Net'].sum().reset_index()

    # Redondear los valores de 'Net' a dos decimales en suma_nets
    suma_nets['Net'] = suma_nets['Net'].round(2)

    # Redondear los valores de 'MONTO' a dos decimales en df_consolidados_filtrados_sorted
    df_consolidados_filtrados_sorted['MONTO'] = df_consolidados_filtrados_sorted['MONTO'].round(2)

    # Función para llenar 'OBS'
    def llenar_obs(row):
        payout_id = row['Payout ID']
        total_net = suma_nets[suma_nets['Payout ID'] == payout_id]['Net'].values
        if len(total_net) == 0:
            return "NP"
        total_net = total_net[0]
        fecha_pago = pd.to_datetime(row['Payout date'], format='%d/%m/%Y')
        monto = df_consolidados_filtrados_sorted[
            (df_consolidados_filtrados_sorted['MONTO'] == total_net) &
            (pd.to_datetime(df_consolidados_filtrados_sorted['FECHA'], format='%d/%m/%Y') > fecha_pago)
        ]
        if not monto.empty:
            return "P"
        return "NP"

    df_pagos_verificada['OBS'] = df_pagos_verificada.apply(llenar_obs, axis=1)

    # Leer la hoja 'SMOOBU BOOKING'
    df_smoobu = pd.read_excel(archivo_excel_original, sheet_name=hoja_smoobu_booking)

    # Asegurarse de que las columnas 'RESERVA' y 'Reference number' tengan el mismo formato (general)
    df_smoobu['RESERVA'] = df_smoobu['RESERVA'].astype(str)
    df_pagos_verificada['Reference number'] = df_pagos_verificada['Reference number'].astype(str)

    # Crear la nueva columna 'PAGO MENOS COMISION'
    df_smoobu['PAGO MENOS COMISION'] = df_smoobu['Precio'] - df_smoobu['Comisión incluida']

    # Redondear la columna 'PAGO MENOS COMISION' a dos decimales
    df_smoobu['PAGO MENOS COMISION'] = df_smoobu['PAGO MENOS COMISION'].round(2)

    # Insertar la nueva columna 'PAGO MENOS COMISION' a la izquierda de 'City tax'
    columnas = df_smoobu.columns.tolist()
    index_city_tax = columnas.index('City tax')
    columnas.insert(index_city_tax, columnas.pop(columnas.index('PAGO MENOS COMISION')))
    df_smoobu = df_smoobu[columnas]

    # Crear la nueva columna 'Payment charge B' al lado derecho de 'PAGO MENOS COMISION'
    index_pago_menos_comision = columnas.index('PAGO MENOS COMISION')
    columnas.insert(index_pago_menos_comision + 1, 'Payment charge B')
    df_smoobu = df_smoobu.reindex(columns=columnas)
    df_smoobu['Payment charge B'] = 0  # Inicializar la columna 'Payment charge B' con valores 0

    # Actualizar la columna 'Payment charge B' con los valores correspondientes de 'Payment charge'
    # Asegurarse de que el 'Reference number' sea único
    df_pagos_verificada_unique = df_pagos_verificada.drop_duplicates(subset='Reference number')

    df_smoobu['Payment charge B'] = df_smoobu['RESERVA'].map(
        df_pagos_verificada_unique.set_index('Reference number')['Payment charge']
    ).fillna(0)

    # Crear la nueva columna 'Unnamed: 11 B' al lado derecho de 'Payment charge B'
    index_payment_charge_b = columnas.index('Payment charge B')
    columnas.insert(index_payment_charge_b + 1, 'Unnamed: 11 B')
    df_smoobu = df_smoobu.reindex(columns=columnas)
    df_smoobu['Unnamed: 11 B'] = 0  # Inicializar la columna 'Unnamed: 11 B' con valores 0

    # Verificar y manejar duplicados en 'Reference number'
    # Agregar una columna de 'Unnamed: 11' desde df_pagos_verificada
    df_smoobu['Unnamed: 11 B'] = df_smoobu['RESERVA'].map(
        df_pagos_verificada.drop_duplicates('Reference number').set_index('Reference number')['Unnamed: 11']
    ).fillna(0)

    # Crear la nueva columna 'NET B' al lado derecho de 'Unnamed: 11 B'
    index_unnamed_11_b = columnas.index('Unnamed: 11 B')
    columnas.insert(index_unnamed_11_b + 1, 'NET B')
    df_smoobu = df_smoobu.reindex(columns=columnas)
    df_smoobu['NET B'] = df_smoobu['PAGO MENOS COMISION'] + df_smoobu['Payment charge B'] + df_smoobu['Unnamed: 11 B']

    # Redondear la columna 'NET B' a dos decimales
    df_smoobu['NET B'] = df_smoobu['NET B'].round(2)

    # Agregar la columna 'OBSERVACION B' al lado derecho de 'RESERVA'
    index_reserva = columnas.index('RESERVA')
    columnas.insert(index_reserva + 1, 'OBSERVACION B')
    df_smoobu = df_smoobu.reindex(columns=columnas)

    # Función para llenar 'OBSERVACION B'
    def observar_b(row):
        reserva = row['RESERVA']
        net_b = row['NET B']
        pago_verificado = df_pagos_verificada[
            (df_pagos_verificada['Reference number'] == reserva) &
            (df_pagos_verificada['Net'] == net_b) &
            (df_pagos_verificada['OBS'] == 'P')
        ]
        if not pago_verificado.empty:
            return "PAGADO"
        return "NO PAGADO"

    df_smoobu['OBSERVACION B'] = df_smoobu.apply(observar_b, axis=1)

    # Escribir los DataFrames en un nuevo archivo Excel con las nuevas hojas
    with pd.ExcelWriter(nuevo_archivo_excel, engine='openpyxl') as writer:
        df_consolidados_filtrados_sorted.to_excel(writer, sheet_name=nueva_hoja_consolidados, index=False)
        df_pagos_verificada.to_excel(writer, sheet_name=nueva_hoja_pagos_booking, index=False)
        df_smoobu.to_excel(writer, sheet_name=nueva_hoja_smoobu_booking, index=False)

    # Cargar el archivo Excel para aplicar el formato condicional
    wb = load_workbook(nuevo_archivo_excel)
    ws_pagos_booking = wb[nueva_hoja_pagos_booking]
    ws_smoobu_booking = wb[nueva_hoja_smoobu_booking]

    # Definir los colores para el formato condicional
    fill_green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Aplicar el formato condicional a la columna 'OBSERVACION' en 'PAGOS BOOKING VERIFICADA'
    obs_col_pagos = ws_pagos_booking['S']  # Asumiendo que la columna 'OBSERVACION' está en la columna S
    for cell in obs_col_pagos:
        if cell.value == "PAGADO":
            cell.fill = fill_green
        elif cell.value == "NO PAGADO":
            cell.fill = fill_red

    # Aplicar el formato condicional a la columna 'OBSERVACION B' en 'SMOOBU BOOKING VERIFICADA'
    obs_col_smoobu = ws_smoobu_booking['AD']  # Asumiendo que la columna 'OBSERVACION B' está en la columna AD
    for cell in obs_col_smoobu:
        if cell.value == "PAGADO":
            cell.fill = fill_green
        elif cell.value == "NO PAGADO":
            cell.fill = fill_red

    # Guardar el archivo con los cambios
    wb.save(nuevo_archivo_excel)

    st.success(f"El nuevo archivo '{nuevo_archivo_excel}' ha sido creado con las hojas '{nueva_hoja_consolidados}', '{nueva_hoja_pagos_booking}', y '{nueva_hoja_smoobu_booking}'.")
    
    #df_final = pd.concat([df_consolidados_filtrados_sorted, df_pagos_verificada, df_smoobu], axis=1)
    #return df_final


    return nuevo_archivo_excel
    #pass

# Interfaz de usuario con Streamlit
def main():
    st.title("Verificación de Pagos Airbnb y Booking")

    st.write("PASO 1:  Carga el archivo Excel original de Airbnb o Booking para comenzar la verificación.")

    uploaded_file = st.file_uploader("Subir archivo Excel", type=["xlsx"])

    if uploaded_file is not None:
        st.write("PASO 2:  Elige la plataforma para la cual deseas realizar la verificación:")
        option = st.selectbox("Selecciona la plataforma", ["Airbnb", "Booking"])

        if st.button("Iniciar verificación"):
            if option == "Airbnb":
                nuevo_archivo_excel=revisar_airbnb(uploaded_file)
                #df_resultante=revisar_airbnb(uploaded_file)
            elif option == "Booking":
                nuevo_archivo_excel=revisar_booking(uploaded_file)
                #df_resultante=revisar_booking(uploaded_file)

            # Guardar el archivo Excel verificado
            #nuevo_archivo_excel = 'nuevo_archivo_verificado.xlsx'
            
            #df_resultante.to_excel(nuevo_archivo_excel, index=False)
            

             # Verificar si nuevo_archivo_excel no es None
            if nuevo_archivo_excel is not None:
                # Verificar si el archivo existe antes de intentar abrirlo
                if os.path.exists(nuevo_archivo_excel):
                    with open(nuevo_archivo_excel, 'rb') as file:
                        btn = st.download_button(
                            label='Descargar archivo Excel verificado',
                            data=file,
                            file_name=nuevo_archivo_excel,
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                        )
                else:
                    st.error("El archivo no se pudo encontrar. Por favor, verifica que la función de revisión haya generado el archivo correctamente.")
            else:
                st.error("La función de revisión no devolvió un nombre de archivo válido. Por favor, verifica las funciones revisar_airbnb y revisar_booking.")

if __name__ == "__main__":
    main()
